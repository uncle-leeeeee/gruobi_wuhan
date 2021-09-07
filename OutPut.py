# -*- coding: utf-8 -*-
"""
Created on Wed Apr  7 16:50:54 2021

@author: 11268
"""

from robust_risk_cost import*
from openpyxl import*
book=Workbook()

#保存选址信息
def output_location():
    sh1 = book.create_sheet("location")
    #输入医院
    i=2
    for g in Transfer:
        sh1.cell(row=i,column=1,value=g)
        i+=1
    #输入情景
    i=2
    for c in scenario:
        sh1.cell(row=1,column=i,value=c)
        i+=1
    
    #输入对应数值
    i=2
    for t in Transfer:
        j=2
        for c in scenario:
            sh1.cell(row=i,column=j,value=O[t].x)
            j+=1
        i+=1

#保存x运量

def output_x():
    sh2=book.create_sheet("x_volume")
    k=1#用于控制情景字符输出位置
    #输出情景
    for c in scenario:
        m=k+1#控制每个情景开始输出数据的位置
        sh2.cell(row=k,column=1,value=(c))
        k+=1
        #输出回收中心
        j=2
        for r in Recyle:
            sh2.cell(row=k-1,column=j,value=(r))
            j+=1
        #输出医院
        i=k
        for g in hospitals:
            sh2.cell(row=i,column=1,value=(g))
            i+=1
            k+=1
        k+=1#每个情景之间多空一行
        #输出对应数据
        
        for g in hospitals:
            j=2
            for r in Recyle:
                sh2.cell(row=m,column=j,value=(sum(x[w,g,r,c].x for w in W)))
                j+=1
            m+=1

#保存y运量
def output_y():
    sh3=book.create_sheet("y_volume")
    k=1#用于控制情景字符输出位置
    #输出情景
    for c in scenario:
        m=k+1#控制每个情景开始输出数据的位置
        sh3.cell(row=k,column=1,value=(c))
        k+=1
        #输出回收中心
        j=2
        for r in Disposal:
            sh3.cell(row=k-1,column=j,value=(r))
            j+=1
        #输出医院
        i=k
        for g in hospitals:
            sh3.cell(row=i,column=1,value=(g))
            i+=1
            k+=1
        k+=1#每个情景之间多空一行
        #输出对应数据
        
        for g in hospitals:
            j=2
            for r in Disposal:
                sh3.cell(row=m,column=j,value=(sum(y[w,g,r,c].x for w in W)))
                j+=1
            m+=1

#保存z运量
def output_z():
    sh4=book.create_sheet("z_volume")
    k=1#用于控制情景字符输出位置
    #输出情景
    for c in scenario:
        m=k+1#控制每个情景开始输出数据的位置
        sh4.cell(row=k,column=1,value=(c))
        k+=1
        #输出回收中心
        j=2
        for r in Transfer:
            sh4.cell(row=k-1,column=j,value=(r))
            j+=1
        #输出医院
        i=k
        for g in hospitals:
            sh4.cell(row=i,column=1,value=(g))
            i+=1
            k+=1
        k+=1#每个情景之间多空一行
        #输出对应数据
        
        for g in hospitals:
            j=2
            for r in Transfer:
                sh4.cell(row=m,column=j,value=(sum(z[w,g,r,c].x for w in W)))
                j+=1
            m+=1

#保存n运量
def output_n():
    sh5=book.create_sheet("n_volume")
    k=1#用于控制情景字符输出位置
    #输出情景
    for c in scenario:
        m=k+1#控制每个情景开始输出数据的位置
        sh5.cell(row=k,column=1,value=(c))
        k+=1
        #输出回收中心
        j=2
        for r in Recyle:
            sh5.cell(row=k-1,column=j,value=(r))
            j+=1
        #输出医院
        i=k
        for g in Transfer:
            sh5.cell(row=i,column=1,value=(g))
            i+=1
            k+=1
        k+=1#每个情景之间多空一行
        #输出对应数据
        
        for g in Transfer:
            j=2
            for r in Recyle:
                sh5.cell(row=m,column=j,value=(sum(n[w,g,r,c].x for w in W)))
                j+=1
            m+=1

#保存m运量
def output_m():
    sh6=book.create_sheet("m_volume")
    k=1#用于控制情景字符输出位置
    #输出情景
    for c in scenario:
        z=k+1#控制每个情景开始输出数据的位置
        sh6.cell(row=k,column=1,value=(c))
        k+=1
        #输出回收中心
        j=2
        for r in Disposal:
            sh6.cell(row=k-1,column=j,value=(r))
            j+=1
        #输出医院
        i=k
        for g in Transfer:
            sh6.cell(row=i,column=1,value=(g))
            i+=1
            k+=1
        k+=1#每个情景之间多空一行
        #输出对应数据
        
        for g in Transfer:
            j=2
            for r in Disposal:
                sh6.cell(row=z,column=j,value=(sum(m[w,g,r,c].x for w in W)))
                j+=1
            z+=1

#保存l1运量
def output_l1():
    sh7=book.create_sheet("l_volume1")
    k=1#用于控制情景字符输出位置
    #输出情景
    for c in scenario:
        z=k+1#控制每个情景开始输出数据的位置
        sh7.cell(row=k,column=1,value=(c))
        k+=1
        #输出回收中心
        j=2
        for r in Disposal:
            sh7.cell(row=k-1,column=j,value=(r))
            j+=1
        #输出医院
        i=k
        for g in Recyle:
            sh7.cell(row=i,column=1,value=(g))
            i+=1
            k+=1
        k+=1#每个情景之间多空一行
        #输出对应数据
        
        for g in Recyle:
            j=2
            for r in Disposal:
                sh7.cell(row=z,column=j,value=(sum(l[w,g,r,'s1',c].x for w in W)))
                j+=1
            z+=1

#保存l2运量
def output_l2():
    sh8=book.create_sheet("l_volume2")
    k=1#用于控制情景字符输出位置
    #输出情景
    for c in scenario:
        z=k+1#控制每个情景开始输出数据的位置
        sh8.cell(row=k,column=1,value=(c))
        k+=1
        #输出回收中心
        j=2
        for r in Disposal:
            sh8.cell(row=k-1,column=j,value=(r))
            j+=1
        #输出医院
        i=k
        for g in Recyle:
            sh8.cell(row=i,column=1,value=(g))
            i+=1
            k+=1
        k+=1#每个情景之间多空一行
        #输出对应数据
        
        for g in Recyle:
            j=2
            for r in Disposal:
                sh8.cell(row=z,column=j,value=(sum(l[w,g,r,'s2',c].x for w in W)))
                j+=1
            z+=1

#输出距离矩阵
def output_GTmatrix():
    sh9 = book.create_sheet("GTmatrix")
    
    i=2
    for g in hospitals:
        sh9.cell(row=i,column=1,value=g)
        i+=1
    #输入情景
    i=2
    for c in Transfer:
        sh9.cell(row=1,column=i,value=c)
        i+=1
    
    #输入对应数值
    i=2
    for t in hospitals:
        j=2
        for c in Transfer:
            sh9.cell(row=i,column=j,value=GTmatrix[t,c])
            j+=1
        i+=1

#输出距离矩阵
def output_GRmatrix():
    sh10 = book.create_sheet("GRmatrix")
    
    i=2
    for g in hospitals:
        sh10.cell(row=i,column=1,value=g)
        i+=1
    #输入情景
    i=2
    for c in Recyle:
        sh10.cell(row=1,column=i,value=c)
        i+=1
    
    #输入对应数值
    i=2
    for t in hospitals:
        j=2
        for c in Recyle:
            sh10.cell(row=i,column=j,value=GRmatrix[t,c])
            j+=1
        i+=1
    
#输出距离矩阵
def output_GDmatrix():
    sh11 = book.create_sheet("GDmatrix")
    
    i=2
    for g in hospitals:
        sh11.cell(row=i,column=1,value=g)
        i+=1
    #输入情景
    i=2
    for c in Disposal:
        sh11.cell(row=1,column=i,value=c)
        i+=1
    
    #输入对应数值
    i=2
    for t in hospitals:
        j=2
        for c in Disposal:
            sh11.cell(row=i,column=j,value=GDmatrix[t,c])
            j+=1
        i+=1

#输出距离矩阵
def output_TRmatrix():
    sh12 = book.create_sheet("TRmatrix")
    
    i=2
    for g in Transfer:
        sh12.cell(row=i,column=1,value=g)
        i+=1
    #输入情景
    i=2
    for c in Recyle:
        sh12.cell(row=1,column=i,value=c)
        i+=1
    
    #输入对应数值
    i=2
    for t in Transfer:
        j=2
        for c in Recyle:
            sh12.cell(row=i,column=j,value=TRmatrix[t,c])
            j+=1
        i+=1

#输出距离矩阵
def output_TDmatrix():
    sh13 = book.create_sheet("TDmatrix")
    
    i=2
    for g in Transfer:
        sh13.cell(row=i,column=1,value=g)
        i+=1
    #输入情景
    i=2
    for c in Disposal:
        sh13.cell(row=1,column=i,value=c)
        i+=1
    
    #输入对应数值
    i=2
    for t in Transfer:
        j=2
        for c in Disposal:
            sh13.cell(row=i,column=j,value=TDmatrix[t,c])
            j+=1
        i+=1

def output_RDmatrix():
    sh14 = book.create_sheet("RDmatrix")
    
    i=2
    for g in Recyle:
        sh14.cell(row=i,column=1,value=g)
        i+=1
    #输入情景
    i=2
    for c in Disposal:
        sh14.cell(row=1,column=i,value=c)
        i+=1
    
    #输入对应数值
    i=2
    for t in Recyle:
        j=2
        for c in Disposal:
            sh14.cell(row=i,column=j,value=RDmatrix[t,c])
            j+=1
        i+=1
        
def output_qy():
    sh15=book.create_sheet("qymatrix")
    k=1#用于控制情景字符输出位置
    #输出情景
    for c in scenario:
        z=k+1#控制每个情景开始输出数据的位置
        sh15.cell(row=k,column=1,value=(c))
        k+=1
        #输出回收中心
        j=2
        for r in Disposal:
            sh15.cell(row=k-1,column=j,value=(r))
            j+=1
        #输出医院
        i=k
        for g in hospitals:
            sh15.cell(row=i,column=1,value=(g))
            i+=1
            k+=1
        k+=1#每个情景之间多空一行
        #输出对应数据
        
        for g in hospitals:
            j=2
            for r in Disposal:
                sh15.cell(row=z,column=j,value=(qy_in[g,r,c].x ))
                j+=1
            z+=1
    
def output_GTnv():
    sh9 = book.create_sheet("GTnv")
    
    i=2
    for g in hospitals:
        sh9.cell(row=i,column=1,value=g)
        i+=1
    #输入情景
    i=2
    for c in Transfer:
        sh9.cell(row=1,column=i,value=c)
        i+=1
    
    #输入对应数值
    i=2
    for t in hospitals:
        j=2
        for c in Transfer:
            sh9.cell(row=i,column=j,value=GT_nv[t,c])
            j+=1
        i+=1
    
def output_GRnv():
    sh10 = book.create_sheet("GRnv")
    
    i=2
    for g in hospitals:
        sh10.cell(row=i,column=1,value=g)
        i+=1
    #输入情景
    i=2
    for c in Recyle:
        sh10.cell(row=1,column=i,value=c)
        i+=1
    
    #输入对应数值
    i=2
    for t in hospitals:
        j=2
        for c in Recyle:
            sh10.cell(row=i,column=j,value=GR_nv[t,c])
            j+=1
        i+=1 
    
def output_GDnv():
    sh11 = book.create_sheet("GDnv")
    
    i=2
    for g in hospitals:
        sh11.cell(row=i,column=1,value=g)
        i+=1
    #输入情景
    i=2
    for c in Disposal:
        sh11.cell(row=1,column=i,value=c)
        i+=1
    
    #输入对应数值
    i=2
    for t in hospitals:
        j=2
        for c in Disposal:
            sh11.cell(row=i,column=j,value=GD_nv[t,c])
            j+=1
        i+=1

def output_TRnv():
    sh12 = book.create_sheet("TRnv")
    
    i=2
    for g in Transfer:
        sh12.cell(row=i,column=1,value=g)
        i+=1
    #输入情景
    i=2
    for c in Recyle:
        sh12.cell(row=1,column=i,value=c)
        i+=1
    
    #输入对应数值
    i=2
    for t in Transfer:
        j=2
        for c in Recyle:
            sh12.cell(row=i,column=j,value=TR_nv[t,c])
            j+=1
        i+=1

def output_TDnv():
    sh13 = book.create_sheet("TDnv")
    
    i=2
    for g in Transfer:
        sh13.cell(row=i,column=1,value=g)
        i+=1
    #输入情景
    i=2
    for c in Disposal:
        sh13.cell(row=1,column=i,value=c)
        i+=1
    
    #输入对应数值
    i=2
    for t in Transfer:
        j=2
        for c in Disposal:
            sh13.cell(row=i,column=j,value=TD_nv[t,c])
            j+=1
        i+=1

def output_RDnv():
    sh14 = book.create_sheet("RDnv")
    
    i=2
    for g in Recyle:
        sh14.cell(row=i,column=1,value=g)
        i+=1
    #输入情景
    i=2
    for c in Disposal:
        sh14.cell(row=1,column=i,value=c)
        i+=1
    
    #输入对应数值
    i=2
    for t in Recyle:
        j=2
        for c in Disposal:
            sh14.cell(row=i,column=j,value=RD_nv[t,c])
            j+=1
        i+=1
    

def output_CREH():
    sh15 = book.create_sheet("CREH")
    #输入总成本
    i=1
    sh15.cell(row=i,column=1,value='Cost')
    i+=1
    sh15.cell(row=i,column=1,value='Risk')
    i+=1
    for c in scenario:
        str1='H['+c+']'
        sh15.cell(row=i,column=1,value=str1)
        i+=1
    for c in scenario:
        str2='E['+c+']'
        sh15.cell(row=i,column=1,value=str2)
        i+=1
    
    i=1
    sh15.cell(row=i,column=2,value=Cost.x)
    i+=1
    sh15.cell(row=i,column=2,value=Risk.x)
    i+=1
    for c in scenario:
        sh15.cell(row=i,column=2,value=H[c].x)
        i+=1
    for c in scenario:
        sh15.cell(row=i,column=2,value=E[c].x)
        i+=1
        
        
        
        
    

    