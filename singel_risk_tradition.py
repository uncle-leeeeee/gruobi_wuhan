# -*- coding: utf-8 -*-
"""
Created on Tue Aug 31 12:12:36 2021

@author: LZH
"""

from data_robust import*
Transfer=['T1','T2','T5']
scenario, p, zeta = gp.multidict({
    'smooth': [1, 1],
})
tau = 0  # 控制偏差的权重系数


#定义model
M=gp.Model("sencond-trial(risk)")

#------------------------------------------------------------------
#定义决策变量
#关于运量
x=M.addVars(W,hospitals,Recyle,scenario,vtype=gp.GRB.CONTINUOUS,name='x')
y=M.addVars(W,hospitals,Disposal,scenario,vtype=gp.GRB.CONTINUOUS,name='y')
z=M.addVars(W,hospitals,Transfer,scenario,vtype=gp.GRB.CONTINUOUS,name='z')
n=M.addVars(W,Transfer,Recyle,scenario,vtype=gp.GRB.CONTINUOUS,name='n')
m=M.addVars(W,Transfer,Disposal,scenario,vtype=gp.GRB.CONTINUOUS,name='m')
l=M.addVars(W,Recyle,Disposal,S,scenario,vtype=gp.GRB.CONTINUOUS,name='l')

#关于每个中心的处理量
r=M.addVars(W,Recyle,S,scenario,vtype=gp.GRB.CONTINUOUS,name='r')
t=M.addVars(W,Transfer,scenario,vtype=gp.GRB.CONTINUOUS,name='t')
d=M.addVars(W,Disposal,S,scenario,vtype=gp.GRB.CONTINUOUS,name='d')

#关于成本与风险目标函数的嵌套变量
E=M.addVars(scenario,vtype=gp.GRB.CONTINUOUS,name='E')
H=M.addVars(scenario,vtype=gp.GRB.CONTINUOUS,name='H')

#两个线性化需要用到的辅助变量
# delta_cost=M.addVars(scenario,vtype=gp.GRB.CONTINUOUS,name='delta_cost')
# delta_risk=M.addVars(scenario,vtype=gp.GRB.CONTINUOUS,name='delta_risk')

#关于风险目标函数，因构建带有指数的目标函数所需的辅助变量
qx_in=M.addVars(hospitals,Recyle,scenario,vtype=gp.GRB.CONTINUOUS,name='qx_in')
qx_out=M.addVars(hospitals,Recyle,scenario,vtype=gp.GRB.CONTINUOUS,name='qx_out')
qy_in=M.addVars(hospitals,Disposal,scenario,vtype=gp.GRB.CONTINUOUS,name='qy_in')
qy_out=M.addVars(hospitals,Disposal,scenario,vtype=gp.GRB.CONTINUOUS,name='qy_out')
qz_in=M.addVars(hospitals,Transfer,scenario,vtype=gp.GRB.CONTINUOUS,name='qz_in')
qz_out=M.addVars(hospitals,Transfer,scenario,vtype=gp.GRB.CONTINUOUS,name='qz_out')
qn_in=M.addVars(Transfer,Recyle,scenario,vtype=gp.GRB.CONTINUOUS,name='qn_in')
qn_out=M.addVars(Transfer,Recyle,scenario,vtype=gp.GRB.CONTINUOUS,name='qn_out')
qm_in=M.addVars(Transfer,Disposal,scenario,vtype=gp.GRB.CONTINUOUS,name='qm_in')
qm_out=M.addVars(Transfer,Disposal,scenario,vtype=gp.GRB.CONTINUOUS,name='qm_out')
ql_in=M.addVars(Recyle,Disposal,scenario,vtype=gp.GRB.CONTINUOUS,name='ql_in')
ql_out=M.addVars(Recyle,Disposal,scenario,vtype=gp.GRB.CONTINUOUS,name='ql_out')
qr=M.addVars(Recyle,scenario,vtype=gp.GRB.CONTINUOUS,name='qr')
qt=M.addVars(Transfer,scenario,vtype=gp.GRB.CONTINUOUS,name='qt')
qd=M.addVars(Disposal,scenario,vtype=gp.GRB.CONTINUOUS,name='qd')

Qx_in=M.addVars(hospitals,Recyle,scenario,vtype=gp.GRB.CONTINUOUS,name='Qx_in')
Qx_out=M.addVars(hospitals,Recyle,scenario,vtype=gp.GRB.CONTINUOUS,name='Qx_out')
Qy_in=M.addVars(hospitals,Disposal,scenario,vtype=gp.GRB.CONTINUOUS,name='Qy_in')
Qy_out=M.addVars(hospitals,Disposal,scenario,vtype=gp.GRB.CONTINUOUS,name='Qy_out')
Qz_in=M.addVars(hospitals,Transfer,scenario,vtype=gp.GRB.CONTINUOUS,name='Qz_in')
Qz_out=M.addVars(hospitals,Transfer,scenario,vtype=gp.GRB.CONTINUOUS,name='Qz_out')
Qn_in=M.addVars(Transfer,Recyle,scenario,vtype=gp.GRB.CONTINUOUS,name='Qn_in')
Qn_out=M.addVars(Transfer,Recyle,scenario,vtype=gp.GRB.CONTINUOUS,name='Qn_out')
Qm_in=M.addVars(Transfer,Disposal,scenario,vtype=gp.GRB.CONTINUOUS,name='Qm_in')
Qm_out=M.addVars(Transfer,Disposal,scenario,vtype=gp.GRB.CONTINUOUS,name='Qm_out')
Ql_in=M.addVars(Recyle,Disposal,scenario,vtype=gp.GRB.CONTINUOUS,name='Ql_in')
Ql_out=M.addVars(Recyle,Disposal,scenario,vtype=gp.GRB.CONTINUOUS,name='Ql_out')
Qr=M.addVars(Recyle,scenario,vtype=gp.GRB.CONTINUOUS,name='Qr')
Qt=M.addVars(Transfer,scenario,vtype=gp.GRB.CONTINUOUS,name='Qt')
Qd=M.addVars(Disposal,scenario,vtype=gp.GRB.CONTINUOUS,name='Qd')

Cost=M.addVar(vtype=gp.GRB.CONTINUOUS,name='Cost')
Location_Risk=M.addVar(vtype=gp.GRB.CONTINUOUS,name='locationRisk')
Trans_Risk=M.addVar(vtype=gp.GRB.CONTINUOUS,name='transRisk')
Location_Cost=M.addVar(vtype=gp.GRB.CONTINUOUS,name='locationCost')
Trans_Cost=M.addVar(vtype=gp.GRB.CONTINUOUS,name='transCost')
#更新模型变量环境
M.update()
#---------------------------------------------------------------
#添加目标函数
#成本最小化原始模型
# M.setObjective((gp.quicksum(TFC[t]*O[t] for t in Transfer)+TranC*(
#                 gp.quicksum(X[w,i,j]*GRmatrix[i,j] for w in W for i in hospitals for j in Recyle)+
#                 gp.quicksum(Y[w,i,j]*GDmatrix[i,j] for w in W for i in hospitals for j in Disposal)+
#                 gp.quicksum(Z[w,i,j]*GTmatrix[i,j] for w in W for i in hospitals for j in Transfer)+
#                 gp.quicksum(N[w,i,j]*TRmatrix[i,j] for w in W for i in Transfer for j in Recyle)+
#                 gp.quicksum(M1[w,i,j]*TDmatrix[i,j] for w in W for i in Transfer for j in Disposal)+
#                 gp.quicksum(L[w,i,j,s]*RDmatrix[i,j] for s in S for w in W for i in Recyle for j in Disposal)
#                 )),sense=gp.GRB.MINIMIZE
#                 )

#成本最小化加鲁棒
M.addConstr(Cost==gp.quicksum(p[c]*E[c] for c in scenario))

#风险最小化原始模型
# M.setObjective(gp.quicksum(x[w,i,j]* GRQmatrix[i,j]*GRmatrix[i,j] for w in W for i in hospitals for j in Recyle)*n_v*lamd/w_v
# +gp.quicksum(x[w,i,j]*GRQ_outmatrix[i,j]*eta*GRmatrix[i,j]*(2*mu_h*GRmatrix[i,j]/w_man-d_length) for w in W for i in hospitals for j in Recyle)
# +gp.quicksum(y[w,i,j]* GDQmatrix[i,j]*GDmatrix[i,j] for w in W for i in hospitals for j in Disposal)*n_v*lamd/w_v
# +gp.quicksum(y[w,i,j]*GDQ_outmatrix[i,j]*eta*GDmatrix[i,j]*(2*mu_h*GDmatrix[i,j]/w_man-d_length) for w in W for i in hospitals for j in Disposal)
# +gp.quicksum(z[w,i,j]* GTQmatrix[i,j]*GTmatrix[i,j] for w in W for i in hospitals for j in Transfer)*n_v*lamd/w_v
# +gp.quicksum(z[w,i,j]*GTQ_outmatrix[i,j]*eta*GTmatrix[i,j]*(2*mu_h*GTmatrix[i,j]/w_man-d_length) for w in W for i in hospitals for j in Transfer)
# +gp.quicksum(n[w,i,j]* TRQmatrix[i,j]*TRmatrix[i,j] for w in W for i in Transfer for j in Recyle)*n_v*lamd/w_v
# +gp.quicksum(n[w,i,j]*TRQ_outmatrix[i,j]*eta*TRmatrix[i,j]*(2*mu_h*TRmatrix[i,j]/w_man-d_length) for w in W for i in Transfer for j in Recyle)
# +gp.quicksum(m[w,i,j]* TDQmatrix[i,j]*TDmatrix[i,j] for w in W for i in Transfer for j in Disposal)*n_v*lamd/w_v
# +gp.quicksum(m[w,i,j]*TDQ_outmatrix[i,j]*eta*TDmatrix[i,j]*(2*mu_h*TDmatrix[i,j]/w_man-d_length) for w in W for i in Transfer for j in Disposal)
# +gp.quicksum(l[w,i,j,s]*RDQmatrix[i,j]*RDmatrix[i,j] for w in W for i in Recyle for j in Disposal for s in S)*n_v*lamd/w_v
# +gp.quicksum(l[w,i,j,s]*RDQ_outmatrix[i,j]*eta*RDmatrix[i,j]*(2*mu_h*RDmatrix[i,j]/w_man-d_length) for w in W for i in Recyle for j in Disposal for s in S)
# +gp.quicksum(t[w,i]*nodeQ*math.pi*eta_node*(r0+mu_h*T)**2 for w in W for i in Transfer)
# +gp.quicksum(d[w,i,s]*nodeQ*math.pi*eta_node*(r0+mu_h*T)**2 for w in W for i in Disposal for s in S)
# +gp.quicksum(r[w,i,s]*nodeQ*math.pi*eta_node*(r0+mu_h*T)**2 for w in W for i in Recyle for s in S)
# ,sense=gp.GRB.MINIMIZE)

#风险最小化加鲁棒
M.setObjective(gp.quicksum(p[c]*H[c] for c in scenario),sense=gp.GRB.MINIMIZE)

#--------------------------------------------------
#添加约束条件

#关于成本目标函数的Ec
M.addConstrs(E[c]==gp.quicksum(TFC[t] for t in Transfer)+TranC*(
gp.quicksum(x[w,i,j,c]*GRmatrix[i,j] for w in W for i in hospitals for j in Recyle)+
gp.quicksum(y[w,i,j,c]*GDmatrix[i,j] for w in W for i in hospitals for j in Disposal)+
gp.quicksum(z[w,i,j,c]*GTmatrix[i,j] for w in W for i in hospitals for j in Transfer)+
gp.quicksum(n[w,i,j,c]*TRmatrix[i,j] for w in W for i in Transfer for j in Recyle)+
gp.quicksum(m[w,i,j,c]*TDmatrix[i,j] for w in W for i in Transfer for j in Disposal)+
gp.quicksum(l[w,i,j,s,c]*RDmatrix[i,j] for s in S for w in W for i in Recyle for j in Disposal)
) for c in scenario)

# M.addConstrs((E[c] - gp.quicksum(p[c]*E[c] for c in scenario)+ 2*delta_cost[c])>=0 for c in scenario)
# M.addConstrs(delta_cost[c]>=0 for c in scenario)

#关于风险目标函数的Hc
M.addConstrs(H[c]==
(gp.quicksum(qx_in[i,j,c]*GRmatrix[i,j]*GR_nv[i,j]*lamd/w_v for i in hospitals for j in Recyle)
+gp.quicksum(qx_out[i,j,c]*eta*GRmatrix[i,j]*(2*mu_h*GRmatrix[i,j]/w_man-d_length) for i in hospitals for j in Recyle)
+gp.quicksum(qy_in[i,j,c]*GDmatrix[i,j]*GD_nv[i,j]*lamd/w_v for i in hospitals for j in Disposal)
+gp.quicksum(qy_out[i,j,c]*eta*GDmatrix[i,j]*(2*mu_h*GDmatrix[i,j]/w_man-d_length) for i in hospitals for j in Disposal)
+gp.quicksum(qz_in[i,j,c]*GTmatrix[i,j]*GT_nv[i,j]*lamd/w_v for i in hospitals for j in Transfer)
+gp.quicksum(qz_out[i,j,c]*eta*GTmatrix[i,j]*(2*mu_h*GTmatrix[i,j]/w_man-d_length) for i in hospitals for j in Transfer)
+gp.quicksum(qn_in[i,j,c]*TRmatrix[i,j]*TR_nv[i,j]*lamd/w_v for i in Transfer for j in Recyle)
+gp.quicksum(qn_out[i,j,c]*eta*TRmatrix[i,j]*(2*mu_h*TRmatrix[i,j]/w_man-d_length) for i in Transfer for j in Recyle)
+gp.quicksum(qm_in[i,j,c]*TDmatrix[i,j]*TD_nv[i,j]*lamd/w_v for i in Transfer for j in Disposal)
+gp.quicksum(qm_out[i,j,c]*eta*TDmatrix[i,j]*(2*mu_h*TDmatrix[i,j]/w_man-d_length) for i in Transfer for j in Disposal)
+gp.quicksum(ql_in[i,j,c]*RDmatrix[i,j]*RD_nv[i,j]*lamd/w_v for i in Recyle for j in Disposal)
+gp.quicksum(ql_out[i,j,c]*eta*RDmatrix[i,j]*(2*mu_h*RDmatrix[i,j]/w_man-d_length) for i in Recyle for j in Disposal)
+gp.quicksum(qr[i,c]*nodeQ*math.pi*eta_node[i]*(r0+mu_h*T)**2 for i in Recyle)
+gp.quicksum(qt[i,c]*nodeQ*math.pi*eta_node[i]*(r0+mu_h*T)**2 for i in Transfer)
+gp.quicksum(qd[i,c]*nodeQ*math.pi*eta_node[i]*(r0+mu_h*T)**2 for i in Disposal))
 for c in scenario
 )

# M.addConstrs((H[c] - gp.quicksum(p[c]*H[c] for c in scenario)+ 2*delta_risk[c])>=0 for c in scenario)
# M.addConstrs(delta_risk[c]>=0 for c in scenario)

#添加Trans_Risk约束，观察总风险当中运输风险的占比
M.addConstr(Trans_Risk==gp.quicksum(qx_in[i,j,'smooth']*GRmatrix[i,j]*GR_nv[i,j]*lamd/w_v for i in hospitals for j in Recyle)
+gp.quicksum(qx_out[i,j,'smooth']*eta*GRmatrix[i,j]*(2*mu_h*GRmatrix[i,j]/w_man-d_length) for i in hospitals for j in Recyle)
+gp.quicksum(qy_in[i,j,'smooth']*GDmatrix[i,j]*GD_nv[i,j]*lamd/w_v for i in hospitals for j in Disposal)
+gp.quicksum(qy_out[i,j,'smooth']*eta*GDmatrix[i,j]*(2*mu_h*GDmatrix[i,j]/w_man-d_length) for i in hospitals for j in Disposal)
+gp.quicksum(qz_in[i,j,'smooth']*GTmatrix[i,j]*GT_nv[i,j]*lamd/w_v for i in hospitals for j in Transfer)
+gp.quicksum(qz_out[i,j,'smooth']*eta*GTmatrix[i,j]*(2*mu_h*GTmatrix[i,j]/w_man-d_length) for i in hospitals for j in Transfer)
+gp.quicksum(qn_in[i,j,'smooth']*TRmatrix[i,j]*TR_nv[i,j]*lamd/w_v for i in Transfer for j in Recyle)
+gp.quicksum(qn_out[i,j,'smooth']*eta*TRmatrix[i,j]*(2*mu_h*TRmatrix[i,j]/w_man-d_length) for i in Transfer for j in Recyle)
+gp.quicksum(qm_in[i,j,'smooth']*TDmatrix[i,j]*TD_nv[i,j]*lamd/w_v for i in Transfer for j in Disposal)
+gp.quicksum(qm_out[i,j,'smooth']*eta*TDmatrix[i,j]*(2*mu_h*TDmatrix[i,j]/w_man-d_length) for i in Transfer for j in Disposal)
+gp.quicksum(ql_in[i,j,'smooth']*RDmatrix[i,j]*RD_nv[i,j]*lamd/w_v for i in Recyle for j in Disposal)
+gp.quicksum(ql_out[i,j,'smooth']*eta*RDmatrix[i,j]*(2*mu_h*RDmatrix[i,j]/w_man-d_length) for i in Recyle for j in Disposal))

#添加Location_Risk，观察总风险当中节点风险的占比
M.addConstr(Location_Risk==gp.quicksum(qr[i,'smooth']*nodeQ*math.pi*eta_node[i]*(r0+mu_h*T)**2 for i in Recyle)
+gp.quicksum(qt[i,'smooth']*nodeQ*math.pi*eta_node[i]*(r0+mu_h*T)**2 for i in Transfer)
+gp.quicksum(qd[i,'smooth']*nodeQ*math.pi*eta_node[i]*(r0+mu_h*T)**2 for i in Disposal))

#添加选址成本，运输成本对应的观察变量
M.addConstr(Location_Cost==gp.quicksum(TFC[t] for t in Transfer))
M.addConstr(Trans_Cost==TranC*(
gp.quicksum(x[w,i,j,'smooth']*GRmatrix[i,j] for w in W for i in hospitals for j in Recyle)+
gp.quicksum(y[w,i,j,'smooth']*GDmatrix[i,j] for w in W for i in hospitals for j in Disposal)+
gp.quicksum(z[w,i,j,'smooth']*GTmatrix[i,j] for w in W for i in hospitals for j in Transfer)+
gp.quicksum(n[w,i,j,'smooth']*TRmatrix[i,j] for w in W for i in Transfer for j in Recyle)+
gp.quicksum(m[w,i,j,'smooth']*TDmatrix[i,j] for w in W for i in Transfer for j in Disposal)+
gp.quicksum(l[w,i,j,s,'smooth']*RDmatrix[i,j] for s in S for w in W for i in Recyle for j in Disposal)
))

#关于qx等，先进行求和约束
M.addConstrs(qx_in[i,j,c]==gp.quicksum(x[w,i,j,c]for w in W) for i in hospitals for j in Recyle for c in scenario)
M.addConstrs(qx_out[i,j,c]==gp.quicksum(x[w,i,j,c]for w in W) for i in hospitals for j in Recyle for c in scenario)
M.addConstrs(qy_in[i,j,c]==gp.quicksum(y[w,i,j,c] for w in W) for i in hospitals for j in Disposal for c in scenario)
M.addConstrs(qy_out[i,j,c]==gp.quicksum(y[w,i,j,c] for w in W) for i in hospitals for j in Disposal for c in scenario)
M.addConstrs(qz_in[i,j,c]==gp.quicksum(z[w,i,j,c] for w in W) for i in hospitals for j in Transfer for c in scenario)
M.addConstrs(qz_out[i,j,c]==gp.quicksum(z[w,i,j,c] for w in W) for i in hospitals for j in Transfer for c in scenario)
M.addConstrs(qn_in[i,j,c]==gp.quicksum(n[w,i,j,c] for w in W) for i in Transfer for j in Recyle for c in scenario)
M.addConstrs(qn_out[i,j,c]==gp.quicksum(n[w,i,j,c] for w in W) for i in Transfer for j in Recyle for c in scenario)
M.addConstrs(qm_in[i,j,c]==gp.quicksum(m[w,i,j,c]for w in W) for i in Transfer for j in Disposal for c in scenario)
M.addConstrs(qm_out[i,j,c]==gp.quicksum(m[w,i,j,c] for w in W) for i in Transfer for j in Disposal for c in scenario)
M.addConstrs(ql_in[i,j,c]==gp.quicksum(l[w,i,j,s,c] for w in W for s in S) for i in Recyle for j in Disposal for c in scenario)
M.addConstrs(ql_out[i,j,c]==gp.quicksum(l[w,i,j,s,c] for w in W for s in S) for i in Recyle for j in Disposal for c in scenario)
M.addConstrs(qr[i,c]==gp.quicksum(r[w,i,s,c] for w in W for s in S) for i in Recyle for c in scenario)
M.addConstrs(qt[i,c]==gp.quicksum(t[w,i,c] for w in W) for i in Transfer for c in scenario)
M.addConstrs(qd[i,c]==gp.quicksum(d[w,i,s,c] for w in W for s in S) for i in Disposal for c in scenario)


#关于Q_x
# for i  in hospitals:
#     for j in Recyle:
#         for c in scenario:
#             string=''.join([i,j,c])
#             M.addGenConstrPow(qx_in[i,j,c],Qx_in[i,j,c],zeta[c],string,"FuncPieces=100")
            
# for i  in hospitals:
#     for j in Recyle:
#         for c in scenario:
#             string=''.join([i,j,c])
#             M.addGenConstrPow(qx_out[i,j,c],Qx_out[i,j,c],zeta[c],string+'_',"FuncPieces=100")


# #Q_y
# for i  in hospitals:
#     for j in Disposal:
#         for c in scenario:
#             string=''.join([i,j,c])
#             M.addGenConstrPow(qy_in[i,j,c],Qy_in[i,j,c],zeta[c],string,"FuncPieces=100")

# for i  in hospitals:
#     for j in Disposal:
#         for c in scenario:
#             string=''.join([i,j,c])
#             M.addGenConstrPow(qy_out[i,j,c],Qy_out[i,j,c],zeta[c],string+'_',"FuncPieces=100")

# #Q_z
# for i  in hospitals:
#     for j in Transfer:
#         for c in scenario:
#             string=''.join([i,j,c])
#             M.addGenConstrPow(qz_in[i,j,c],Qz_in[i,j,c],zeta[c],string,"FuncPieces=100")

# for i  in hospitals:
#     for j in Transfer:
#         for c in scenario:
#             string=''.join([i,j,c])
#             M.addGenConstrPow(qz_out[i,j,c],Qz_out[i,j,c],zeta[c],string+'_',"FuncPieces=100")

# #Q_n
# for i  in Transfer:
#     for j in Recyle:
#         for c in scenario:
#             string=''.join([i,j,c])
#             M.addGenConstrPow(qn_in[i,j,c],Qn_in[i,j,c],zeta[c],string,"FuncPieces=100")

# for i  in Transfer:
#     for j in Recyle:
#         for c in scenario:
#             string=''.join([i,j,c])
#             M.addGenConstrPow(qn_out[i,j,c],Qn_out[i,j,c],zeta[c],string+'_',"FuncPieces=100")

# #Q_m
# for i  in Transfer:
#     for j in Disposal:
#         for c in scenario:
#             string=''.join([i,j,c])
#             M.addGenConstrPow(qm_in[i,j,c],Qm_in[i,j,c],zeta[c],string,"FuncPieces=100")

# for i  in Transfer:
#     for j in Disposal:
#         for c in scenario:
#             string=''.join([i,j,c])
#             M.addGenConstrPow(qm_out[i,j,c],Qm_out[i,j,c],zeta[c],string+'_',"FuncPieces=100")

# #Q_l
# for i  in Recyle:
#     for j in Disposal:
#         for c in scenario:
#             string=''.join([i,j,c])
#             M.addGenConstrPow(ql_in[i,j,c],Ql_in[i,j,c],zeta[c],string,"FuncPieces=100")

# for i  in Recyle:
#     for j in Disposal:
#         for c in scenario:
#             string=''.join([i,j,c])
#             M.addGenConstrPow(ql_out[i,j,c],Ql_out[i,j,c],zeta[c],string+'_',"FuncPieces=100")

# #Q_r
# for i in Recyle:
#     for c in scenario:
#         string=''.join([i,c])
#         M.addGenConstrPow(qr[i,c],Qr[i,c],zeta[c],string,"FuncPieces=100")
        
# #Q_t
# for i in Transfer:
#     for c in scenario:
#         string=''.join([i,c])
#         M.addGenConstrPow(qt[i,c],Qt[i,c],zeta[c],string,"FuncPieces=100")
        
# #Q_d
# for i in Disposal:
#     for c in scenario:
#         string=''.join([i,c])
#         M.addGenConstrPow(qd[i,c],Qd[i,c],zeta[c],string,"FuncPieces=100")

# for i  in hospitals:
#     for j in Recyle:
#         for c in scenario:
#             string=''.join([i,j,c])
#             M.addGenConstrPow(qx_in[i,j,c],Qx_in[i,j,c],zeta[c],string,"FuncPieces=1000")
            
# for i  in hospitals:
#     for j in Recyle:
#         for c in scenario:
#             string=''.join([i,j,c])
#             M.addGenConstrPow(qx_out[i,j,c],Qx_out[i,j,c],zeta[c],string+'_',"FuncPieces=1000")


# #Q_y
# for i  in hospitals:
#     for j in Disposal:
#         for c in scenario:
#             string=''.join([i,j,c])
#             M.addGenConstrPow(qy_in[i,j,c],Qy_in[i,j,c],zeta[c],string,"FuncPieces=1000")

# for i  in hospitals:
#     for j in Disposal:
#         for c in scenario:
#             string=''.join([i,j,c])
#             M.addGenConstrPow(qy_out[i,j,c],Qy_out[i,j,c],zeta[c],string+'_',"FuncPieces=1000")

# #Q_z
# for i  in hospitals:
#     for j in Transfer:
#         for c in scenario:
#             string=''.join([i,j,c])
#             M.addGenConstrPow(qz_in[i,j,c],Qz_in[i,j,c],zeta[c],string,"FuncPieces=1000")

# for i  in hospitals:
#     for j in Transfer:
#         for c in scenario:
#             string=''.join([i,j,c])
#             M.addGenConstrPow(qz_out[i,j,c],Qz_out[i,j,c],zeta[c],string+'_',"FuncPieces=1000")

# #Q_n
# for i  in Transfer:
#     for j in Recyle:
#         for c in scenario:
#             string=''.join([i,j,c])
#             M.addGenConstrPow(qn_in[i,j,c],Qn_in[i,j,c],zeta[c],string,"FuncPieces=1000")

# for i  in Transfer:
#     for j in Recyle:
#         for c in scenario:
#             string=''.join([i,j,c])
#             M.addGenConstrPow(qn_out[i,j,c],Qn_out[i,j,c],zeta[c],string+'_',"FuncPieces=1000")

# #Q_m
# for i  in Transfer:
#     for j in Disposal:
#         for c in scenario:
#             string=''.join([i,j,c])
#             M.addGenConstrPow(qm_in[i,j,c],Qm_in[i,j,c],zeta[c],string,"FuncPieces=1000")

# for i  in Transfer:
#     for j in Disposal:
#         for c in scenario:
#             string=''.join([i,j,c])
#             M.addGenConstrPow(qm_out[i,j,c],Qm_out[i,j,c],zeta[c],string+'_',"FuncPieces=1000")

# #Q_l
# for i  in Recyle:
#     for j in Disposal:
#         for c in scenario:
#             string=''.join([i,j,c])
#             M.addGenConstrPow(ql_in[i,j,c],Ql_in[i,j,c],zeta[c],string,"FuncPieces=1000")

# for i  in Recyle:
#     for j in Disposal:
#         for c in scenario:
#             string=''.join([i,j,c])
#             M.addGenConstrPow(ql_out[i,j,c],Ql_out[i,j,c],zeta[c],string+'_',"FuncPieces=1000")

# #Q_r
# for i in Recyle:
#     for c in scenario:
#         string=''.join([i,c])
#         M.addGenConstrPow(qr[i,c],Qr[i,c],zeta[c],string,"FuncPieces=1000")
        
# #Q_t
# for i in Transfer:
#     for c in scenario:
#         string=''.join([i,c])
#         M.addGenConstrPow(qt[i,c],Qt[i,c],zeta[c],string,"FuncPieces=1000")
        
# #Q_d
# for i in Disposal:
#     for c in scenario:
#         string=''.join([i,c])
#         M.addGenConstrPow(qd[i,c],Qd[i,c],zeta[c],string,"FuncPieces=1000")





#第一阶段
M.addConstrs(alpha*generations[i,w]==gp.quicksum(y[w,i,j,c] for j in Disposal) for w in W for i in hospitals for c in scenario)
M.addConstrs(beta*generations[i,w]==gp.quicksum(x[w,i,j,c] for j in Recyle) for w in W for i in hospitals for c in scenario)
M.addConstrs((1-alpha-beta)*generations[i,w]==gp.quicksum(z[w,i,j,c] for j in Transfer) for w in W for i in hospitals for c in scenario)
M.addConstrs(r[w,j,'s1',c]==gp.quicksum(x[w,i,j,c] for i in hospitals ) for w in W for j in Recyle for c in scenario)
M.addConstrs(t[w,i,c]==gp.quicksum(z[w,j,i,c] for j in hospitals) for w in W for i in Transfer for c in scenario)
M.addConstrs((1-delta)*r[w,i,'s1',c]==gp.quicksum(l[w,i,j,'s1',c] for j in Disposal) for i in Recyle for w in W  for c in scenario)
M.addConstrs((d[w,i,'s1',c]==gp.quicksum(l[w,j,i,'s1',c] for j in Recyle)+gp.quicksum(y[w,j,i,c] for j in hospitals)) for i in Disposal for w in W for c in scenario)

#第二阶段
M.addConstrs(t[w,i,c]*gamma==gp.quicksum(n[w,i,j,c]for j in Recyle) for w in W for i in Transfer for c in scenario)
M.addConstrs((1-gamma)*t[w,i,c]==gp.quicksum(m[w,i,j,c] for j in Disposal) for w in W for i in Transfer for c in scenario)
M.addConstrs(r[w,i,'s2',c]==gp.quicksum(n[w,j,i,c] for j in Transfer) for w in W for i in Recyle for c in scenario)
M.addConstrs((1-delta)*r[w,i,'s2',c]==gp.quicksum(l[w,i,j,'s2',c] for j in Disposal) for w in W for i in Recyle for c in scenario)
M.addConstrs(d[w,i,'s2',c]==gp.quicksum(l[w,j,i,'s2',c] for j in Recyle ) + gp.quicksum(m[w,j,i,c] for j in Transfer) for w in W  for i in Disposal for c in scenario)

#能力约束
M.addConstrs(gp.quicksum(r[w,i,s,c] for w in W)<=RC[i] for s in S for i in Recyle for c in scenario)
M.addConstrs(gp.quicksum(t[w,i,c] for w in W)<=TC[i] for i in Transfer for c in scenario)
M.addConstrs(gp.quicksum(d[w,i,s,c] for w in W ) <= DC[i] for i in Disposal for s in S for c in scenario)

#逻辑约束
# M.addConstrs(x[w,i,j,c] <=MAX*X[w,i,j,c] for w in W for i in hospitals for j in Recyle for c in scenario)
# M.addConstrs(y[w,i,j,c] <=MAX*Y[w,i,j,c] for w in W for i in hospitals for j in Disposal for c in scenario)
# M.addConstrs(z[w,i,j,c] <=MAX*Z[w,i,j,c] for w in W for i in hospitals for j in Transfer for c in scenario)
# M.addConstrs(n[w,i,j,c] <=MAX*N[w,i,j,c] for w in W for i in Transfer for j in Recyle for c in scenario)
# M.addConstrs(m[w,i,j,c] <=MAX*M1[w,i,j,c] for w in W for i in Transfer for j in Disposal for c in scenario)
# M.addConstrs(l[w,i,j,s,c] <=MAX*L[w,i,j,s,c] for w in W for i in Recyle for j in Disposal for s in S for c in scenario)
M.addConstrs(x[w,i,j,c]>=0 for w in W for i in hospitals for j in Recyle for c in scenario)
M.addConstrs(y[w,i,j,c]>=0 for w in W for i in hospitals for j in Disposal for c in scenario)
M.addConstrs(z[w,i,j,c]>=0 for w in W for i in hospitals for j in Transfer for c in scenario)
M.addConstrs(n[w,i,j,c]>=0 for w in W for i in Transfer for j in Recyle for c in scenario)
M.addConstrs(m[w,i,j,c]>=0 for w in W for i in Transfer for j in Disposal for c in scenario)
M.addConstrs(l[w,i,j,s,c]>=0 for w in W for i in Recyle for j in Disposal for s in S for c in scenario)

M.update()
M.Params.MIPGap=0.02
M.Params.timeLimit=1200.0

M.optimize()

if M.status==gp.GRB.Status.OPTIMAL:
    print('Optimal solution found')


M.write(r'50_risk_5.lp')
M.write(r'50_risk_5.attr')
M.write(r'50_risk_5.json')


from openpyxl import*
book=Workbook()

#保存选址信息
# def output_location():
#     sh1 = book.create_sheet("location")
#     #输入医院
#     i=2
#     for g in Transfer:
#         sh1.cell(row=i,column=1,value=g)
#         i+=1
#     #输入情景
#     i=2
#     for c in scenario:
#         sh1.cell(row=1,column=i,value=c)
#         i+=1
    
#     #输入对应数值
#     i=2
#     for t in Transfer:
#         j=2
#         for c in scenario:
#             sh1.cell(row=i,column=j,value=O[t].x)
#             j+=1
#         i+=1

#保存x运量

def output_x():
    sh2=book.create_sheet("x_volume")
    k=1#用于控制情景字符输出位置
    #输出情景
    for c in scenario:
        m=k+1#控制每个情景开始输出数据的位置
        sh2.cell(row=k,column=1,value=(c))
        k+=1
        #输出回收中心
        j=2
        for r in Recyle:
            sh2.cell(row=k-1,column=j,value=(r))
            j+=1
        #输出医院
        i=k
        for g in hospitals:
            sh2.cell(row=i,column=1,value=(g))
            i+=1
            k+=1
        k+=1#每个情景之间多空一行
        #输出对应数据
        
        for g in hospitals:
            j=2
            for r in Recyle:
                sh2.cell(row=m,column=j,value=(sum(x[w,g,r,c].x for w in W)))
                j+=1
            m+=1

#保存y运量
def output_y():
    sh3=book.create_sheet("y_volume")
    k=1#用于控制情景字符输出位置
    #输出情景
    for c in scenario:
        m=k+1#控制每个情景开始输出数据的位置
        sh3.cell(row=k,column=1,value=(c))
        k+=1
        #输出回收中心
        j=2
        for r in Disposal:
            sh3.cell(row=k-1,column=j,value=(r))
            j+=1
        #输出医院
        i=k
        for g in hospitals:
            sh3.cell(row=i,column=1,value=(g))
            i+=1
            k+=1
        k+=1#每个情景之间多空一行
        #输出对应数据
        
        for g in hospitals:
            j=2
            for r in Disposal:
                sh3.cell(row=m,column=j,value=(sum(y[w,g,r,c].x for w in W)))
                j+=1
            m+=1

#保存z运量
def output_z():
    sh4=book.create_sheet("z_volume")
    k=1#用于控制情景字符输出位置
    #输出情景
    for c in scenario:
        m=k+1#控制每个情景开始输出数据的位置
        sh4.cell(row=k,column=1,value=(c))
        k+=1
        #输出回收中心
        j=2
        for r in Transfer:
            sh4.cell(row=k-1,column=j,value=(r))
            j+=1
        #输出医院
        i=k
        for g in hospitals:
            sh4.cell(row=i,column=1,value=(g))
            i+=1
            k+=1
        k+=1#每个情景之间多空一行
        #输出对应数据
        
        for g in hospitals:
            j=2
            for r in Transfer:
                sh4.cell(row=m,column=j,value=(sum(z[w,g,r,c].x for w in W)))
                j+=1
            m+=1

#保存n运量
def output_n():
    sh5=book.create_sheet("n_volume")
    k=1#用于控制情景字符输出位置
    #输出情景
    for c in scenario:
        m=k+1#控制每个情景开始输出数据的位置
        sh5.cell(row=k,column=1,value=(c))
        k+=1
        #输出回收中心
        j=2
        for r in Recyle:
            sh5.cell(row=k-1,column=j,value=(r))
            j+=1
        #输出医院
        i=k
        for g in Transfer:
            sh5.cell(row=i,column=1,value=(g))
            i+=1
            k+=1
        k+=1#每个情景之间多空一行
        #输出对应数据
        
        for g in Transfer:
            j=2
            for r in Recyle:
                sh5.cell(row=m,column=j,value=(sum(n[w,g,r,c].x for w in W)))
                j+=1
            m+=1

#保存m运量
def output_m():
    sh6=book.create_sheet("m_volume")
    k=1#用于控制情景字符输出位置
    #输出情景
    for c in scenario:
        z=k+1#控制每个情景开始输出数据的位置
        sh6.cell(row=k,column=1,value=(c))
        k+=1
        #输出回收中心
        j=2
        for r in Disposal:
            sh6.cell(row=k-1,column=j,value=(r))
            j+=1
        #输出医院
        i=k
        for g in Transfer:
            sh6.cell(row=i,column=1,value=(g))
            i+=1
            k+=1
        k+=1#每个情景之间多空一行
        #输出对应数据
        
        for g in Transfer:
            j=2
            for r in Disposal:
                sh6.cell(row=z,column=j,value=(sum(m[w,g,r,c].x for w in W)))
                j+=1
            z+=1

#保存l1运量
def output_l1():
    sh7=book.create_sheet("l_volume1")
    k=1#用于控制情景字符输出位置
    #输出情景
    for c in scenario:
        z=k+1#控制每个情景开始输出数据的位置
        sh7.cell(row=k,column=1,value=(c))
        k+=1
        #输出回收中心
        j=2
        for r in Disposal:
            sh7.cell(row=k-1,column=j,value=(r))
            j+=1
        #输出医院
        i=k
        for g in Recyle:
            sh7.cell(row=i,column=1,value=(g))
            i+=1
            k+=1
        k+=1#每个情景之间多空一行
        #输出对应数据
        
        for g in Recyle:
            j=2
            for r in Disposal:
                sh7.cell(row=z,column=j,value=(sum(l[w,g,r,'s1',c].x for w in W)))
                j+=1
            z+=1

#保存l2运量
def output_l2():
    sh8=book.create_sheet("l_volume2")
    k=1#用于控制情景字符输出位置
    #输出情景
    for c in scenario:
        z=k+1#控制每个情景开始输出数据的位置
        sh8.cell(row=k,column=1,value=(c))
        k+=1
        #输出回收中心
        j=2
        for r in Disposal:
            sh8.cell(row=k-1,column=j,value=(r))
            j+=1
        #输出医院
        i=k
        for g in Recyle:
            sh8.cell(row=i,column=1,value=(g))
            i+=1
            k+=1
        k+=1#每个情景之间多空一行
        #输出对应数据
        
        for g in Recyle:
            j=2
            for r in Disposal:
                sh8.cell(row=z,column=j,value=(sum(l[w,g,r,'s2',c].x for w in W)))
                j+=1
            z+=1

#输出距离矩阵
def output_GTmatrix():
    sh9 = book.create_sheet("GTmatrix")
    
    i=2
    for g in hospitals:
        sh9.cell(row=i,column=1,value=g)
        i+=1
    #输入情景
    i=2
    for c in Transfer:
        sh9.cell(row=1,column=i,value=c)
        i+=1
    
    #输入对应数值
    i=2
    for t in hospitals:
        j=2
        for c in Transfer:
            sh9.cell(row=i,column=j,value=GTmatrix[t,c])
            j+=1
        i+=1
        
def output_GTnv():
    sh9 = book.create_sheet("GTnv")
    
    i=2
    for g in hospitals:
        sh9.cell(row=i,column=1,value=g)
        i+=1
    #输入情景
    i=2
    for c in Transfer:
        sh9.cell(row=1,column=i,value=c)
        i+=1
    
    #输入对应数值
    i=2
    for t in hospitals:
        j=2
        for c in Transfer:
            sh9.cell(row=i,column=j,value=GT_nv[t,c])
            j+=1
        i+=1

def output_CREH():
    sh15 = book.create_sheet("CREH")
    #输入总成本
    i=1
    sh15.cell(row=i,column=1,value='Cost')
    i+=1
    sh15.cell(row=i,column=1,value='Risk')
    i+=1
    for c in scenario:
        str1='H['+c+']'
        sh15.cell(row=i,column=1,value=str1)
        i+=1
    for c in scenario:
        str2='E['+c+']'
        sh15.cell(row=i,column=1,value=str2)
        i+=1
    sh15.cell(row=i,column=1,value='Trans_Cost')
    i+=1
    sh15.cell(row=i,column=1,value='Trans_Risk')
    i+=1
    sh15.cell(row=i,column=1,value='Location_Cost')    
    i+=1
    sh15.cell(row=i,column=1,value='Location_Risk')    
    i+=1
    
    i=1
    sh15.cell(row=i,column=2,value=Cost.x)
    i+=1
    sh15.cell(row=i,column=2,value=M.objval)
    i+=1
    for c in scenario:
        sh15.cell(row=i,column=2,value=H[c].x)
        i+=1
    for c in scenario:
        sh15.cell(row=i,column=2,value=E[c].x)
        i+=1
    sh15.cell(row=i,column=2,value=Trans_Cost.x)
    i+=1
    sh15.cell(row=i,column=2,value=Trans_Risk.x)
    i+=1
    sh15.cell(row=i,column=2,value=Location_Cost.x)    
    i+=1
    sh15.cell(row=i,column=2,value=Location_Risk.x)    
    i+=1
    

# output_location()
output_x()
output_y()
output_z()
output_m()
output_n()
output_l1()
output_l2()
output_CREH()
# str4='D:\\gruobi练习\\中规模双目标问题求解(50) _wuhan\\\''+'输出数据双目标50_'+str(u)+'.xlsx'
str4='D:\\gruobi练习\\中规模双目标问题求解(50) _wuhan\\\''+'分阶段_5'+'.xlsx'
book.save(str4)
        






